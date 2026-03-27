#!/usr/bin/env python3
"""
Reports Merger — Combine all CSV reports into one Excel file
=============================================================
Every time you run bangalore_ads_report.py, a new CSV is saved
in the reports_archive/ folder. This script combines ALL of them
into one Excel file with:
  - Sheet 1: All data combined (no duplicates)
  - Sheet 2: Summary by search category
  - Sheet 3: Top rated companies
  - One extra sheet per report date

Usage:
    pip3 install openpyxl --break-system-packages
    python3 merge_reports.py
    open master_report.xlsx
"""

import csv
import os
import sys
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system("pip3 install openpyxl --break-system-packages")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

REPORTS_DIR  = "reports_archive"
OUTPUT_EXCEL = "master_report.xlsx"

# ── Styles ────────────────────────────────────────────────────
BLUE_FILL   = PatternFill("solid", fgColor="1E5AA0")
GRAY_FILL   = PatternFill("solid", fgColor="EFF4FB")
GREEN_FILL  = PatternFill("solid", fgColor="E9F7EF")
YELLOW_FILL = PatternFill("solid", fgColor="FEF9E7")
WHITE_FONT  = Font(color="FFFFFF", bold=True, size=11)
BOLD_FONT   = Font(bold=True, size=11)
HEADER_FONT = Font(bold=True, size=12, color="1E5AA0")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = BLUE_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_row(ws, row_num, num_cols, alt=False):
    fill = GRAY_FILL if alt else PatternFill("solid", fgColor="FFFFFF")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=False)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Load all CSV files ────────────────────────────────────────
def load_all_reports():
    if not os.path.exists(REPORTS_DIR):
        print(f"Error: {REPORTS_DIR}/ folder not found.")
        print("Run python3 bangalore_ads_report.py first to generate reports.")
        sys.exit(1)

    all_files = sorted([f for f in os.listdir(REPORTS_DIR) if f.endswith(".csv")])
    if not all_files:
        print(f"No CSV files found in {REPORTS_DIR}/")
        sys.exit(1)

    print(f"\n  Found {len(all_files)} report(s) in {REPORTS_DIR}/")
    all_rows = []
    file_data = {}

    for fname in all_files:
        fpath = os.path.join(REPORTS_DIR, fname)
        rows = []
        with open(fpath, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                row["_source_file"] = fname
                row["_fetched_date"] = fname[:8]  # YYYYMMDD
                rows.append(row)
        file_data[fname] = rows
        all_rows.extend(rows)
        print(f"    {fname}: {len(rows)} companies")

    return all_rows, file_data


# ── Deduplicate ───────────────────────────────────────────────
def deduplicate(rows):
    seen = {}
    for row in rows:
        key = row.get("place_id") or row.get("name", "")
        if key not in seen:
            seen[key] = row
    return list(seen.values())


# ── Sheet 1: All companies ────────────────────────────────────
def write_all_sheet(wb, all_unique):
    ws = wb.active
    ws.title = "All Companies"
    ws.row_dimensions[1].height = 35

    headers = ["#", "Company Name", "Address", "Rating", "Total Ratings",
               "Phone", "Website", "Category", "Status", "Fetched Date"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = h
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    sorted_rows = sorted(all_unique, key=lambda x: float(x.get("rating") or 0), reverse=True)
    for i, row in enumerate(sorted_rows, 1):
        data = [
            i,
            row.get("name", ""),
            row.get("address", ""),
            row.get("rating", ""),
            row.get("total_ratings", ""),
            row.get("phone", ""),
            row.get("website", ""),
            row.get("category", row.get("_source_file", "")),
            row.get("business_status", ""),
            row.get("_fetched_date", ""),
        ]
        for col, val in enumerate(data, 1):
            ws.cell(row=i+1, column=col).value = val
        style_data_row(ws, i+1, len(headers), alt=(i % 2 == 0))

    set_col_widths(ws, [5, 35, 40, 8, 14, 16, 35, 35, 12, 14])
    print(f"  Sheet 1 'All Companies': {len(sorted_rows)} unique companies")


# ── Sheet 2: Summary ──────────────────────────────────────────
def write_summary_sheet(wb, all_unique, file_data):
    ws = wb.create_sheet("Summary by Category")
    ws.row_dimensions[1].height = 30

    # Title
    ws.merge_cells("A1:F1")
    title = ws.cell(row=1, column=1)
    title.value = f"Report Summary — Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title.font = HEADER_FONT
    title.alignment = Alignment(horizontal="center")

    headers = ["Report File", "Date", "Companies", "Avg Rating", "With Phone", "With Website"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col).value = h
    style_header_row(ws, 2, len(headers))

    for i, (fname, rows) in enumerate(file_data.items(), 1):
        rated = [r for r in rows if r.get("rating")]
        avg = round(sum(float(r["rating"]) for r in rated) / len(rated), 1) if rated else "N/A"
        data = [
            fname,
            fname[:8],
            len(rows),
            avg,
            sum(1 for r in rows if r.get("phone")),
            sum(1 for r in rows if r.get("website")),
        ]
        for col, val in enumerate(data, 1):
            ws.cell(row=i+2, column=col).value = val
        style_data_row(ws, i+2, len(headers), alt=(i % 2 == 0))

    # Totals row
    tot_row = len(file_data) + 3
    ws.cell(row=tot_row, column=1).value = "TOTAL (unique)"
    ws.cell(row=tot_row, column=3).value = len(all_unique)
    for col in range(1, 7):
        ws.cell(row=tot_row, column=col).fill = GREEN_FILL
        ws.cell(row=tot_row, column=col).font = BOLD_FONT

    set_col_widths(ws, [45, 12, 12, 12, 12, 14])
    print(f"  Sheet 2 'Summary by Category': {len(file_data)} reports")


# ── Sheet 3: Top Rated ────────────────────────────────────────
def write_top_rated_sheet(wb, all_unique):
    ws = wb.create_sheet("Top Rated ⭐")
    ws.row_dimensions[1].height = 30

    top = [r for r in all_unique if r.get("rating") and float(r["rating"]) >= 4.5]
    top = sorted(top, key=lambda x: float(x.get("rating") or 0), reverse=True)

    ws.merge_cells("A1:F1")
    title = ws.cell(row=1, column=1)
    title.value = f"Top Rated Companies (≥4.5 ⭐) — {len(top)} companies"
    title.font = HEADER_FONT
    title.alignment = Alignment(horizontal="center")

    headers = ["#", "Company Name", "Rating", "Phone", "Website", "Address"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col).value = h
    style_header_row(ws, 2, len(headers))

    for i, row in enumerate(top, 1):
        data = [i, row.get("name",""), row.get("rating",""),
                row.get("phone",""), row.get("website",""), row.get("address","")]
        for col, val in enumerate(data, 1):
            ws.cell(row=i+2, column=col).value = val
        style_data_row(ws, i+2, len(headers), alt=(i % 2 == 0))

    set_col_widths(ws, [5, 35, 8, 16, 35, 40])
    print(f"  Sheet 3 'Top Rated': {len(top)} companies")


# ── Per-report sheets ─────────────────────────────────────────
def write_per_report_sheets(wb, file_data):
    for fname, rows in file_data.items():
        sheet_name = fname[:25].replace("_", " ").strip()
        ws = wb.create_sheet(sheet_name)

        headers = ["#", "Name", "Rating", "Phone", "Website", "Address", "Status"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col).value = h
        style_header_row(ws, 1, len(headers))

        sorted_rows = sorted(rows, key=lambda x: float(x.get("rating") or 0), reverse=True)
        for i, row in enumerate(sorted_rows, 1):
            data = [i, row.get("name",""), row.get("rating",""),
                    row.get("phone",""), row.get("website",""),
                    row.get("address",""), row.get("business_status","")]
            for col, val in enumerate(data, 1):
                ws.cell(row=i+1, column=col).value = val
            style_data_row(ws, i+1, len(headers), alt=(i % 2 == 0))

        set_col_widths(ws, [5, 32, 8, 16, 32, 38, 12])


# ── Main ─────────────────────────────────────────────────────
def main():
    print("\n" + "=" * 60)
    print("  Reports Merger — Building Master Excel File")
    print("=" * 60)

    all_rows, file_data = load_all_reports()
    all_unique = deduplicate(all_rows)
    print(f"\n  Total rows loaded  : {len(all_rows)}")
    print(f"  Unique companies   : {len(all_unique)}")
    print(f"\n  Building Excel sheets...")

    wb = openpyxl.Workbook()
    write_all_sheet(wb, all_unique)
    write_summary_sheet(wb, all_unique, file_data)
    write_top_rated_sheet(wb, all_unique)
    write_per_report_sheets(wb, file_data)

    wb.save(OUTPUT_EXCEL)

    print(f"\n  ✓ Master Excel saved: {OUTPUT_EXCEL}")
    print(f"  Sheets: All Companies | Summary | Top Rated | + {len(file_data)} report sheets")
    print(f"\n  Open it: open {OUTPUT_EXCEL}")
    print("=" * 60)


if __name__ == "__main__":
    main()
